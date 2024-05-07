from openpyxl import load_workbook, workbook
import openpyxl
import pandas as pd


def read_excel_data_pandas(file_name,sheet):
    test_data = pd.read_excel(file_name, sheet_name=sheet)
    for index, rows in test_data.iterrows():
        username = test_data['username']
        password = test_data['password']
        expected = test_data['exp']

def read_excel_data(file_name, sheet):
    data_list =[]
    try:
        wb = load_workbook(filename=file_name)
        sh = wb[sheet]
        rows = sh.max_row
        cols = sh.max_column
        for i in range(2, rows+1):
            row=[]
            for j in range(1, cols+1):
                row.append(sh.cell(row=i, column=j).value)
            data_list.append(row)
    except Exception as e:
        print(e.with_traceback())
    return data_list



def getRowCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return(sheet.max_row)

def getColumnCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return(sheet.max_column)

def readData(file,sheetName,rownum,columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(row=rownum, column=columnno).value

def writeData(file,sheetName,rownum,columnno,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(row=rownum, column=columnno).value = data
    workbook.save(file)