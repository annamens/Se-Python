#Python selenium
import pytest
import  requests
import pymysql
from selenium import webdriver
from selenium.common import ElementNotVisibleException, ElementNotSelectableException
from selenium.webdriver.chrome.options import Options
driver = webdriver.Chrome()
chrome_options = Options
remote_driver = webdriver.Remote(command_executor='host', options=chrome_options)
chrome_options.add_argument('--headless')

#waitsc
#implicit wait
driver.implicitly_wait(10)
#explicit wait
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
element = WebDriverWait(driver, 10).until(EC.visibility_of_element_located(By.XPATH,""))
#Fluent wait
wait = WebDriverWait(driver, 10, poll_frequency=1, ignored_exceptions=[ElementNotVisibleException, ElementNotSelectableException])
element = wait.until(EC.element_to_be_clickable((By.XPATH, "//div")))
#hard waits
import time
time.sleep(10) # hard waits?


driver.get("url")
driver.refresh()
driver.backward()
driver.forward()
driver.find_element(By.XPATH,"xpath")
driver.find_element(By.CSS,"").send_keys("text")
#alerts
alerts = driver.switch_to.alert
alerts.send_keys("text")
alerts.accept()
alerts.dismiss()
#Action chains
from selenium.webdriver.common.action_chains import ActionChains
ac = ActionChains(driver)
ac.move_to_element("el").perform() #to hover
ac.context_click("el").perform()
ac.double_click("el").perform()
ac.drag_and_drop("src_el", "trgt_el").perform()
ac.click_and_hold("el")
ac.release()
#select
from selenium.webdriver.support.select import Select
drop_down = Select("dropdown_el")
drop_down.select_by_index(2)
drop_down.select_by_value("value")
drop_down.selct_by_visible_text("text")
#window handles
current_window= driver.current_window_handle
windows = driver.window_handles[-1] #gives latest window/tab
window = driver.switch_to_window(driver.window_handles[1])
#DB
import pymysql
# import mysql.connector
# conn1= mysql.connector.connect()
with pymysql.connect(
    hostname="",
    username="",
    password="",
    database=""
    ) as conn:

    cursor = conn.cursor()
    cursor.execute("query")

#screenshot
ss = driver.save_screen_shot("ss.png")
#JavaScript
driver.execute_script("script code")
#frames
driver.switch_to.frame("el")
driver.switch_to.default_content()
#file actions, using context manager to open and close files or connections
with open("filename",'a') as file:
    file.write("python-selenium")
#Excel
from openpyxl import load_workbook
wb = load_workbook("file.xlsx")
data = wb.active
for row in data.iter_rows(min_row=2, values_only=True):
  username, password= row
  # using pandas
  import pandas as pd
  #Read test data from Excel file
  test_data = pd.read_excel('testdata.xlsx', sheet_name='sheet1')
  def add(num1, num2):
      return num1 + num2
  # Iterate over rows of test data
  for index, row in test_data.iterrows():
      num1 = row['num1']
      num2 = row['num2']
      expected_result = row['expected_result']
      result = add(num1, num2)
      if result == expected_result:
          print(f"Test case passed: {num1} + {num2} = {result}")
      else:
          print(f"Test case failed: {num1} + {num2} = {result}, expected {expected_result}")
#logging
import logging
logging.basicConfig(filename="filename",
                    format="format",
                    datefmt="")
logger=logging.getLogger()
logger.setLevel(logging.INFO)

#read_properties
import configparser
config=configparser.ConfigParser(),
config.read("filename")
config.get('info','username')

driver.add_cookie("cookie")
cookies=driver.get_cookies()
driver.delete_cookie("cookie")
'''
#cookies
cookie = {
name:"xyz",
value: 'value'
}
#fixtures
@pytest.fixture
fixtures are functions used to provide data, set pre conditions or set up
needed for the test, they are executed before test methods or functions
#markers , custom markers are defined in pytest.ini
@pytest.mark.slow
@pytest.mark.parameters(1,"string")
Markers are used to categorize the tests or to add metadata to tests
they allow us to add certain attributes or porperties to tests

#hooks
Hooks in pytest are functions that allows to customize or extend the behaviour of test scripts
pytest_sessionstart(session)

#decorators
decorators are a powerful feature that allows you to modify or extend the behavior of functions or methods.
#monkey patching
it is to modify a function at run time by assigning new function to old function
#requests for API testing
GET: to retrieve the data from the server
PUT: to update an existing resource/ create new if one doesn't exist
PATCH: update partial data
POST: to create a new resource
DELETE: delete an existing resource
#HTTP status codes
2xx- success
200-OK  -->GET
201-created -->POST
204-no content  -->DELETE
3xx- redirection
301-moved permanently
302-found
304-not modified
4xx-client error
400-Bad request
402-unauthorized
404-not found
5xx-server error
500-internal server error
502-bad gateway
503-service unavailable
'''
def test_get_user(user_id):
  response = requests.get("url", headers="headers")
  user=response.json()

def test_create_user():
  data={
      "name":"john",
      "username":"johndoe",
      "email":"john@gmaul.com"
  }
  response = requests.post("", json=data)
  created_user=response.json()
def test_delete_user():
    response = requests.delete("")